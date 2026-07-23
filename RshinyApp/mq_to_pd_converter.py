#!/usr/bin/env python3
"""
mq_to_pd_converter.py
Konvertiert MaxQuant-Output in ein ProteomeDiscoverer-kompatibles Excel-Format.

Benoetigt: pandas, numpy, openpyxl
Installation: pip install pandas openpyxl

Usage:
    python mq_to_pd_converter.py \
        --modpep  modificationSpecificPeptides.txt \
        --phospho "Phospho (STY) Sites.txt" \
        [--peptides peptides.txt]        # optional: liefert exakte Peptid-Positionen
        [--output  YYYYMMDD_mq_converted.xlsx]
        [--keep-reverse]
        [--keep-contaminants]
"""

import argparse
import re
import os
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import WriteOnlyCell


# ──────────────────────────────────────────────────────────────────────────────
# Konvertierungs-Logik
# ──────────────────────────────────────────────────────────────────────────────

def build_mods_in_master(site_ids_series, phos_idx):
    """'Modifications in Master Proteins' → 'NxPhospho [S13(92); T17(100)]'"""
    def one(s):
        if not s or pd.isna(s):
            return None
        ids = [x.strip() for x in str(s).split(";") if x.strip()]
        sites = []
        for sid in ids:
            if sid not in phos_idx.index:
                continue
            row = phos_idx.loc[sid]
            aa       = str(row["Amino acid"])
            pos      = int(str(row["Positions within proteins"]).split(";")[0])
            prob_pct = int(round(float(row["Localization prob"]) * 100))
            sites.append(f"{aa}{pos}({prob_pct})")
        return f"{len(sites)}xPhospho [{'; '.join(sites)}]" if sites else None
    return site_ids_series.apply(one)


def build_positions_from_peptides_txt(peptide_id_series, pep_idx):
    """'Positions in Master Proteins' direkt aus peptides.txt → '[start-end]'"""
    def one(pid):
        pid = str(pid).strip() if not pd.isna(pid) else ""
        if not pid or pid not in pep_idx.index:
            return None
        row = pep_idx.loc[pid]
        if pd.isna(row["Start position"]):
            return None
        return f"[{int(float(row['Start position']))}-{int(float(row['End position']))}]"
    return peptide_id_series.apply(one)


def build_positions_from_phospho(site_ids_series, seqs_series, phos_idx):
    """Fallback: Positionen aus Phospho_STY berechnen (weniger präzise)."""
    def one(args):
        site_ids_str, seq = args
        if not site_ids_str or pd.isna(site_ids_str) or not seq or pd.isna(seq):
            return None
        ids = [x.strip() for x in str(site_ids_str).split(";") if x.strip()]
        if not ids or ids[0] not in phos_idx.index:
            return None
        row      = phos_idx.loc[ids[0]]
        prot_pos = int(str(row["Positions within proteins"]).split(";")[0])
        pep_pos  = int(row["Position in peptide"])
        pep_start = prot_pos - pep_pos + 1
        pep_end   = pep_start + len(str(seq)) - 1
        return f"[{pep_start}-{pep_end}]"
    return pd.Series(list(map(one, zip(site_ids_series, seqs_series))),
                     index=site_ids_series.index)


def build_modifications_col(modpep):
    """'Modifications' → 'NxAcetyl [N-Term]; NxOxidation [M]'"""
    acetyl    = pd.to_numeric(modpep["Acetyl (Protein N-term)"], errors="coerce").fillna(0).astype(int)
    oxidation = pd.to_numeric(modpep["Oxidation (M)"],           errors="coerce").fillna(0).astype(int)
    def make(a, o):
        parts = []
        if a > 0: parts.append(f"{a}xAcetyl [N-Term]")
        if o > 0: parts.append(f"{o}xOxidation [M]")
        return "; ".join(parts)
    return pd.Series([make(a, o) for a, o in zip(acetyl, oxidation)], index=modpep.index)


# ──────────────────────────────────────────────────────────────────────────────
# Excel-Schreiblogik (write_only für Performance)
# ──────────────────────────────────────────────────────────────────────────────

DARK_BLUE = "2E4057"

COL_WIDTHS = {
    "Peptide ID": 16, "Confidence": 10,
    "Sequence": 34, "Annotated Sequence": 34,
    "Modifications": 30, "# Protein Groups": 10, "# Proteins": 10, "# PSMs": 8,
    "Master Protein Accessions": 22, "Positions in Master Proteins": 16,
    "Modifications in Master Proteins": 38, "Master Protein Descriptions": 30,
    "Protein Accessions": 30, "# Missed Cleavages": 14, "Theo. MH+ [Da]": 14,
    "Contaminant": 12, "Species": 12, "Gene name": 14,
}


def write_excel(out, stat, outfile, sample_names):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Peptides")

    h_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    h_fill  = PatternFill("solid", fgColor=DARK_BLUE)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = list(out.columns)

    # Column widths
    for col_idx, h in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        if h in COL_WIDTHS:                    ws.column_dimensions[letter].width = COL_WIDTHS[h]
        elif h and h.startswith("Sample_"):    ws.column_dimensions[letter].width = 14
        elif h and h.startswith("MQ_"):        ws.column_dimensions[letter].width = 20
        else:                                   ws.column_dimensions[letter].width = 14

    ws.row_dimensions[1].height = 42

    # Header row
    hrow = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_align
        hrow.append(c)
    ws.append(hrow)

    # Data rows
    for i, (_, row) in enumerate(out.iterrows()):
        vals = []
        for v in row:
            if isinstance(v, float) and np.isnan(v):     vals.append(None)
            elif isinstance(v, (np.integer,)):             vals.append(int(v))
            elif isinstance(v, (np.floating,)):            vals.append(float(v) if not np.isnan(v) else None)
            elif v is pd.NA or v is None:                  vals.append(None)
            else:                                          vals.append(v)
        ws.append(vals)
        if (i + 1) % 5000 == 0:
            print(f"  {i+1}/{len(out)} Zeilen geschrieben...")

    # Statistik-Sheet
    ws2 = wb.create_sheet("Statistik")
    for i, (k, v) in enumerate(zip(stat["Feld"], stat["Wert"])):
        if i == 0:
            ck = WriteOnlyCell(ws2, value="Feld")
            ck.font = h_font; ck.fill = h_fill
            cv = WriteOnlyCell(ws2, value="Wert")
            cv.font = h_font; cv.fill = h_fill
            ws2.append([ck, cv])
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 70

    wb.save(outfile)

    # Post-save: add freeze + autofilter (requires non-write-only)
    wb2 = load_workbook(outfile)
    ws3 = wb2["Peptides"]
    ws3.freeze_panes = "E2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws3.sheet_properties.tabColor = DARK_BLUE
    wb2["Statistik"].sheet_properties.tabColor = "4CAF50"
    wb2.save(outfile)


# ──────────────────────────────────────────────────────────────────────────────
# Haupt-Pipeline
# ──────────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="MaxQuant → ProteomeDiscoverer Excel Konverter")
    p.add_argument("--modpep",             required=True,  help="modificationSpecificPeptides.txt")
    p.add_argument("--phospho",            required=True,  help="Phospho (STY) Sites.txt")
    p.add_argument("--peptides",           default=None,   help="peptides.txt (optional, verbessert Positionen)")
    p.add_argument("--output",             default=None,   help="Ausgabedatei (.xlsx)")
    p.add_argument("--samples",            default=None,   help="Komma-getrennte Experiment-Nummern")
    p.add_argument("--keep-reverse",       action="store_true")
    p.add_argument("--keep-contaminants",  action="store_true")
    return p.parse_args()


def read_mq_tsv(path, label="file"):
    """
    Liest MaxQuant TSV-Dateien robust: probiert UTF-8, dann Latin-1 (cp1252).
    MaxQuant auf Windows erzeugt oft Latin-1-kodierte Dateien.
    """
    for enc in ("utf-8", "latin-1", "cp1252", "utf-16"):
        try:
            df = pd.read_csv(path, sep="\t", dtype=str,
                             low_memory=False, encoding=enc)
            if enc != "utf-8":
                print(f"  (Hinweis: {label} wurde als {enc} eingelesen)")
            return df
        except UnicodeDecodeError:
            continue
    raise ValueError(
        f"Konnte {label} mit keinem der getesteten Encodings lesen "
        "(utf-8, latin-1, cp1252, utf-16). Bitte Datei prüfen."
    )


def main():
    args = parse_args()

    print("Lese modificationSpecificPeptides.txt ...")
    modpep = read_mq_tsv(args.modpep, "modificationSpecificPeptides.txt")
    print(f"  {len(modpep)} Zeilen")

    print("Lese Phospho (STY) Sites.txt ...")
    phospho = read_mq_tsv(args.phospho, "Phospho (STY) Sites.txt")
    print(f"  {len(phospho)} Zeilen")

    pep_idx = None
    if args.peptides:
        print("Lese peptides.txt ...")
        pep = read_mq_tsv(args.peptides, "peptides.txt")
        pep_idx = pep.set_index("id")[["Start position", "End position"]].copy()
        print(f"  {len(pep)} Zeilen — Positionen werden direkt übernommen")

    # Filter
    if not args.keep_reverse:
        n = len(modpep)
        modpep = modpep[modpep["Reverse"].fillna("") != "+"].reset_index(drop=True)
        print(f"  Reverse-Hits entfernt: {n - len(modpep)}")
    if not args.keep_contaminants:
        n = len(modpep)
        modpep = modpep[modpep["Potential contaminant"].fillna("") != "+"].reset_index(drop=True)
        print(f"  Kontaminanten entfernt: {n - len(modpep)}")
    print(f"  Verbleibende Peptide: {len(modpep)}")

    phos_idx = phospho.set_index("id")

    # Intensity-Spalten erkennen (numerisch: "Intensity 1" oder benannt: "Intensity A2")
    numeric_int = sorted([c for c in modpep.columns if re.match(r"^Intensity \d+$", c)],
                         key=lambda x: int(x.split()[-1]))
    named_int   = sorted([c for c in modpep.columns if re.match(r"^Intensity [A-Za-z]", c)])

    if args.samples:
        sel = [s.strip() for s in args.samples.split(",")]
        int_cols     = [f"Intensity {n}" for n in sel if f"Intensity {n}" in modpep.columns]
        sample_names = [f"Sample_{n}"    for n in sel if f"Intensity {n}" in modpep.columns]
    elif numeric_int:
        int_cols     = numeric_int
        sample_names = [f"Sample_{c.split()[-1]}" for c in int_cols]
    else:
        int_cols     = named_int
        sample_names = [c.replace("Intensity ", "Sample_") for c in int_cols]

    print(f"  {len(int_cols)} Sample-Spalten erkannt")

    # Berechnungen
    print("Berechne Modifications in Master Proteins ...")
    site_ids       = modpep["Phospho (STY) site IDs"].fillna("")
    mods_in_master = build_mods_in_master(site_ids, phos_idx)

    print("Berechne Positions in Master Proteins ...")
    if pep_idx is not None:
        positions = build_positions_from_peptides_txt(modpep["Peptide ID"], pep_idx)
    else:
        positions = build_positions_from_phospho(site_ids, modpep["Sequence"], phos_idx)

    print("Baue Modifications-Spalte ...")
    modifications = build_modifications_col(modpep)

    lead_prots = modpep["Proteins"].str.split(";").str[0].str.strip()
    n_proteins = modpep["Proteins"].str.split(";").apply(len)

    intensities = modpep[int_cols].apply(pd.to_numeric, errors="coerce")
    intensities  = intensities.where(intensities != 0, other=np.nan)
    intensities.columns = sample_names

    print("Baue Output-DataFrame ...")
    out = pd.DataFrame({
        "Peptide ID":                       "peptide_" + modpep.index.astype(str),
        "Confidence":                       "High",
        "Sequence":                         modpep["Sequence"].values,
        "Annotated Sequence":               modpep["Sequence"].values,
        "Modifications":                    modifications.values,
        "# Protein Groups":                 1,
        "# Proteins":                       n_proteins.values,
        "# PSMs":                           pd.NA,
        "Master Protein Accessions":        lead_prots.values,
        "Positions in Master Proteins":     positions.values,
        "Modifications in Master Proteins": mods_in_master.values,
        "Master Protein Descriptions":      modpep["Protein Names"].values,
        "Protein Accessions":               modpep["Proteins"].values,
        "# Missed Cleavages":               pd.to_numeric(modpep["Missed cleavages"], errors="coerce"),
        "Theo. MH+ [Da]":                   pd.to_numeric(modpep["Mass"], errors="coerce"),
        "Contaminant":                      (modpep["Potential contaminant"] == "+").map({True: "True", False: "False"}),
        "Species":                          pd.NA,
        "Gene name":                        modpep["Gene Names"].values,
    })
    out = pd.concat([out, intensities.reset_index(drop=True)], axis=1)

    # MQ-Zusatzspalten
    for col_out, col_in, numeric in [
        ("MQ_Score",             "Score",                   True),
        ("MQ_PEP",               "PEP",                     True),
        ("MQ_Reverse",           "Reverse",                 False),
        ("MQ_id",                "id",                      False),
        ("MQ_PhosphoSiteIDs",    "Phospho (STY) site IDs",  False),
        ("MQ_ProteinGroupID",    "Protein Groups",          False),
        ("MQ_UniqueGroups",      "Unique (Groups)",         False),
        ("MQ_UniqueProteins",    "Unique (Proteins)",       False),
        ("MQ_Phospho_count",     "Phospho (STY)",           True),
        ("MQ_Oxidation_count",   "Oxidation (M)",           True),
        ("MQ_Acetyl_count",      "Acetyl (Protein N-term)", True),
        ("MQ_Modifications_raw", "Modifications",           False),
    ]:
        if col_in in modpep.columns:
            v = modpep[col_in]
            out[col_out] = pd.to_numeric(v, errors="coerce").values if numeric else v.values

    # Peptide ID aus peptides.txt ebenfalls behalten
    if "Peptide ID" in modpep.columns:
        out["MQ_PeptideID"] = modpep["Peptide ID"].values

    # Statistik
    stat = pd.DataFrame({
        "Feld": [
            "Datum", "Peptide gesamt",
            "mit Phospho", "mit Acetyl (N-Term)", "mit Oxidation (M)", "Sample-Spalten",
            "Positions-Quelle",
            "", "Hinweis: Master Protein Accessions",
            "Hinweis: Positions in Master Proteins",
            "Hinweis: Modifications",
            "Hinweis: Modifications in Master Proteins",
            "Verwendung",
        ],
        "Wert": [
            str(date.today()), len(out),
            int(out["Modifications in Master Proteins"].notna().sum()),
            int(out["Modifications"].str.contains("Acetyl",    na=False).sum()),
            int(out["Modifications"].str.contains("Oxidation", na=False).sum()),
            len(sample_names),
            "peptides.txt (direkt)" if pep_idx is not None else "Phospho_STY (berechnet)",
            "",
            "Erste UniProt AC des Leading Proteins",
            "Aus peptides.txt wenn verfügbar, sonst aus Phospho_STY berechnet",
            "PD-Format: NxAcetyl [N-Term]; NxOxidation [M]. Phospho NICHT hier.",
            "NxPhospho [AA{absPos}({prob%}); ...]. Aus Phospho_STY aggregiert.",
            "run_proforma_pipeline('<diese_datei>.xlsx') in core_calculation.R",
        ],
    })

    outfile = args.output or f"{date.today().strftime('%Y%m%d')}_mq_converted.xlsx"
    print(f"Schreibe Excel: {outfile} ...")
    write_excel(out, stat, outfile, sample_names)

    size_mb = os.path.getsize(outfile) / 1024 / 1024
    print(f"FERTIG → {outfile}  ({len(out)} Zeilen, {len(out.columns)} Spalten, {size_mb:.1f} MB)")


if __name__ == "__main__":
    main()
